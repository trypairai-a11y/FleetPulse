"""
Report generation service for FleetPulse.

Supports 6 report types:
  - attendance      : Attendance records with driver name, date, status, late_minutes
  - orders          : Captured orders with driver, platform, order_ref, amount (KWD)
  - performance     : Driver performance — name, composite score, order count, attendance rate
  - maintenance     : Maintenance records with vehicle, category, cost (KWD), date
  - cash            : Cash records with driver, type, amount (KWD), status, date
  - fleet_overview  : Tenant-wide summary statistics

Supported output formats:  pdf  |  excel  |  csv

Entry point:
    bytes_, filename, content_type = await generate_report(
        report_type, format, date_from, date_to, filters, db
    )
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy import Integer, case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attendance import AttendanceRecord
from app.models.cash import CashRecord
from app.models.driver import Driver
from app.models.order import CapturedOrder
from app.models.vehicle import MaintenanceRecord, Vehicle
from app.models.ai import AIScore

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

KWD_PLACES = 3  # always render monetary values to 3 decimal places

CONTENT_TYPES = {
    "pdf": "application/pdf",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
}

EXTENSIONS = {
    "pdf": "pdf",
    "excel": "xlsx",
    "csv": "csv",
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


async def generate_report(
    report_type: str,
    format: str,  # noqa: A002  (shadows builtin, kept for spec compatibility)
    date_from: date | None,
    date_to: date | None,
    filters: dict[str, Any] | None,
    db: AsyncSession,
) -> tuple[bytes, str, str]:
    """
    Generate a report and return ``(file_bytes, filename, content_type)``.

    Parameters
    ----------
    report_type:
        One of: attendance, orders, performance, maintenance, cash, fleet_overview
    format:
        One of: pdf, excel, csv
    date_from / date_to:
        Optional inclusive date range filter.
    filters:
        Optional extra key/value filters (e.g. ``{"platform": "talabat"}``).
    db:
        Active async SQLAlchemy session (tenant RLS context must already be set
        by the caller via ``SET LOCAL app.current_tenant_id = '...'``).
    """
    filters = filters or {}

    # Dispatch data fetching
    _fetchers = {
        "attendance": _fetch_attendance,
        "orders": _fetch_orders,
        "performance": _fetch_performance,
        "maintenance": _fetch_maintenance,
        "cash": _fetch_cash,
        "fleet_overview": _fetch_fleet_overview,
    }
    if report_type not in _fetchers:
        raise ValueError(f"Unknown report type: {report_type!r}")
    if format not in CONTENT_TYPES:
        raise ValueError(f"Unknown format: {format!r}")

    headers, rows, title = await _fetchers[report_type](db, date_from, date_to, filters)

    # Dispatch rendering
    if format == "pdf":
        file_bytes = _render_pdf(title, headers, rows)
    elif format == "excel":
        file_bytes = _render_excel(title, headers, rows)
    else:  # csv
        file_bytes = _render_csv(headers, rows)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"fleetpulse_{report_type}_{timestamp}.{EXTENSIONS[format]}"
    content_type = CONTENT_TYPES[format]

    return file_bytes, filename, content_type


# ---------------------------------------------------------------------------
# Data fetchers — each returns (headers: list[str], rows: list[list], title: str)
# ---------------------------------------------------------------------------


async def _fetch_attendance(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """Attendance records joined with driver names."""
    stmt = (
        select(
            Driver.name.label("driver_name"),
            Driver.employee_id.label("employee_id"),
            AttendanceRecord.date.label("date"),
            AttendanceRecord.status.label("status"),
            AttendanceRecord.late_minutes.label("late_minutes"),
            AttendanceRecord.source.label("source"),
        )
        .join(Driver, AttendanceRecord.driver_id == Driver.id)
        .order_by(AttendanceRecord.date.desc(), Driver.name)
    )

    stmt = _apply_date_filter(stmt, AttendanceRecord.date, date_from, date_to)

    if filters.get("status"):
        stmt = stmt.where(AttendanceRecord.status == filters["status"])
    if filters.get("driver_id"):
        stmt = stmt.where(AttendanceRecord.driver_id == uuid.UUID(str(filters["driver_id"])))

    result = await db.execute(stmt)
    records = result.all()

    headers = ["Driver Name", "Employee ID", "Date", "Status", "Late Minutes", "Source"]
    rows = [
        [
            r.driver_name,
            r.employee_id or "",
            str(r.date),
            r.status,
            str(r.late_minutes or 0),
            r.source,
        ]
        for r in records
    ]
    return headers, rows, "Attendance Report"


async def _fetch_orders(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """Captured orders joined with driver names."""
    stmt = (
        select(
            Driver.name.label("driver_name"),
            Driver.employee_id.label("employee_id"),
            CapturedOrder.platform.label("platform"),
            CapturedOrder.order_ref.label("order_ref"),
            CapturedOrder.amount.label("amount"),
            CapturedOrder.status.label("status"),
            CapturedOrder.captured_at.label("captured_at"),
        )
        .join(Driver, CapturedOrder.driver_id == Driver.id)
        .order_by(CapturedOrder.captured_at.desc())
    )

    if date_from or date_to:
        stmt = _apply_datetime_date_filter(stmt, CapturedOrder.captured_at, date_from, date_to)

    if filters.get("platform"):
        stmt = stmt.where(CapturedOrder.platform == filters["platform"])
    if filters.get("driver_id"):
        stmt = stmt.where(CapturedOrder.driver_id == uuid.UUID(str(filters["driver_id"])))
    if filters.get("status"):
        stmt = stmt.where(CapturedOrder.status == filters["status"])

    result = await db.execute(stmt)
    records = result.all()

    headers = ["Driver Name", "Employee ID", "Platform", "Order Ref", "Amount (KWD)", "Status", "Captured At"]
    rows = [
        [
            r.driver_name,
            r.employee_id or "",
            r.platform,
            r.order_ref or "",
            _fmt_kwd(r.amount),
            r.status,
            _fmt_datetime(r.captured_at),
        ]
        for r in records
    ]
    return headers, rows, "Orders Report"


async def _fetch_performance(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """
    Driver performance report.

    Aggregates per driver over the date range:
      - Average composite AI score (from ai_scores table)
      - Total order count
      - Attendance rate (present + late) / total records
    """
    # --- AI scores sub-query: avg composite score per driver ---
    score_stmt = (
        select(
            AIScore.driver_id.label("driver_id"),
            func.avg(AIScore.composite_score).label("avg_score"),
        )
        .group_by(AIScore.driver_id)
    )
    score_stmt = _apply_date_filter(score_stmt, AIScore.date, date_from, date_to)
    score_sub = score_stmt.subquery()

    # --- Orders sub-query: count per driver ---
    order_stmt = (
        select(
            CapturedOrder.driver_id.label("driver_id"),
            func.count(CapturedOrder.id).label("order_count"),
        )
        .group_by(CapturedOrder.driver_id)
    )
    if date_from or date_to:
        order_stmt = _apply_datetime_date_filter(
            order_stmt, CapturedOrder.captured_at, date_from, date_to
        )
    order_sub = order_stmt.subquery()

    # --- Attendance sub-query: total and present/late count per driver ---
    att_total_stmt = (
        select(
            AttendanceRecord.driver_id.label("driver_id"),
            func.count(AttendanceRecord.id).label("total_records"),
            func.sum(
                case(
                    (AttendanceRecord.status.in_(["present", "late"]), 1),
                    else_=0,
                )
            ).label("present_count"),
        )
        .group_by(AttendanceRecord.driver_id)
    )
    att_total_stmt = _apply_date_filter(att_total_stmt, AttendanceRecord.date, date_from, date_to)
    att_sub = att_total_stmt.subquery()

    # --- Main query: join drivers with aggregates ---
    stmt = (
        select(
            Driver.name.label("driver_name"),
            Driver.employee_id.label("employee_id"),
            Driver.platform.label("platform"),
            Driver.status.label("driver_status"),
            score_sub.c.avg_score.label("avg_score"),
            order_sub.c.order_count.label("order_count"),
            att_sub.c.total_records.label("total_records"),
            att_sub.c.present_count.label("present_count"),
        )
        .outerjoin(score_sub, Driver.id == score_sub.c.driver_id)
        .outerjoin(order_sub, Driver.id == order_sub.c.driver_id)
        .outerjoin(att_sub, Driver.id == att_sub.c.driver_id)
        .where(Driver.is_active.is_(True))
        .order_by(score_sub.c.avg_score.desc().nullslast(), Driver.name)
    )

    if filters.get("platform"):
        stmt = stmt.where(Driver.platform == filters["platform"])
    if filters.get("driver_id"):
        stmt = stmt.where(Driver.id == uuid.UUID(str(filters["driver_id"])))

    result = await db.execute(stmt)
    records = result.all()

    headers = [
        "Driver Name",
        "Employee ID",
        "Platform",
        "Status",
        "Avg AI Score",
        "Total Orders",
        "Attendance Rate",
        "Present Days",
        "Total Recorded Days",
    ]
    rows = []
    for r in records:
        total = r.total_records or 0
        present = r.present_count or 0
        att_rate = f"{(present / total * 100):.1f}%" if total > 0 else "N/A"
        avg_score = f"{float(r.avg_score):.2f}" if r.avg_score is not None else "N/A"
        rows.append([
            r.driver_name,
            r.employee_id or "",
            r.platform or "",
            r.driver_status,
            avg_score,
            str(r.order_count or 0),
            att_rate,
            str(present),
            str(total),
        ])
    return headers, rows, "Driver Performance Report"


async def _fetch_maintenance(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """Maintenance records joined with vehicle plate/make/model."""
    stmt = (
        select(
            Vehicle.plate_number.label("plate_number"),
            Vehicle.make.label("make"),
            Vehicle.model.label("model"),
            Vehicle.vehicle_type.label("vehicle_type"),
            MaintenanceRecord.date.label("date"),
            MaintenanceRecord.category.label("category"),
            MaintenanceRecord.type.label("maint_type"),
            MaintenanceRecord.cost.label("cost"),
            MaintenanceRecord.status.label("status"),
            MaintenanceRecord.vendor.label("vendor"),
            MaintenanceRecord.description.label("description"),
        )
        .join(Vehicle, MaintenanceRecord.vehicle_id == Vehicle.id)
        .order_by(MaintenanceRecord.date.desc())
    )

    stmt = _apply_date_filter(stmt, MaintenanceRecord.date, date_from, date_to)

    if filters.get("category"):
        stmt = stmt.where(MaintenanceRecord.category == filters["category"])
    if filters.get("status"):
        stmt = stmt.where(MaintenanceRecord.status == filters["status"])
    if filters.get("vehicle_id"):
        stmt = stmt.where(MaintenanceRecord.vehicle_id == uuid.UUID(str(filters["vehicle_id"])))

    result = await db.execute(stmt)
    records = result.all()

    headers = [
        "Plate Number", "Make", "Model", "Type",
        "Date", "Category", "Maintenance Type",
        "Cost (KWD)", "Status", "Vendor", "Description",
    ]
    rows = [
        [
            r.plate_number,
            r.make or "",
            r.model or "",
            r.vehicle_type,
            str(r.date),
            r.category,
            r.maint_type,
            _fmt_kwd(r.cost),
            r.status,
            r.vendor or "",
            r.description or "",
        ]
        for r in records
    ]
    return headers, rows, "Maintenance Report"


async def _fetch_cash(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """Cash records joined with driver names."""
    stmt = (
        select(
            Driver.name.label("driver_name"),
            Driver.employee_id.label("employee_id"),
            CashRecord.date.label("date"),
            CashRecord.record_type.label("record_type"),
            CashRecord.amount.label("amount"),
            CashRecord.status.label("status"),
            CashRecord.deposit_location.label("deposit_location"),
            CashRecord.reference_number.label("reference_number"),
        )
        .join(Driver, CashRecord.driver_id == Driver.id)
        .order_by(CashRecord.date.desc(), Driver.name)
    )

    stmt = _apply_date_filter(stmt, CashRecord.date, date_from, date_to)

    if filters.get("record_type"):
        stmt = stmt.where(CashRecord.record_type == filters["record_type"])
    if filters.get("status"):
        stmt = stmt.where(CashRecord.status == filters["status"])
    if filters.get("driver_id"):
        stmt = stmt.where(CashRecord.driver_id == uuid.UUID(str(filters["driver_id"])))

    result = await db.execute(stmt)
    records = result.all()

    headers = [
        "Driver Name", "Employee ID", "Date",
        "Type", "Amount (KWD)", "Status",
        "Deposit Location", "Reference Number",
    ]
    rows = [
        [
            r.driver_name,
            r.employee_id or "",
            str(r.date),
            r.record_type,
            _fmt_kwd(r.amount),
            r.status,
            r.deposit_location or "",
            r.reference_number or "",
        ]
        for r in records
    ]
    return headers, rows, "Cash Records Report"


async def _fetch_fleet_overview(
    db: AsyncSession,
    date_from: date | None,
    date_to: date | None,
    filters: dict,
) -> tuple[list[str], list[list], str]:
    """Tenant-wide fleet summary statistics."""
    # Total active drivers
    r_drivers = await db.execute(
        select(func.count(Driver.id)).where(Driver.is_active.is_(True))
    )
    total_drivers = r_drivers.scalar() or 0

    # Drivers by platform
    r_platform = await db.execute(
        select(Driver.platform, func.count(Driver.id))
        .where(Driver.is_active.is_(True))
        .group_by(Driver.platform)
        .order_by(Driver.platform)
    )
    platform_counts = r_platform.all()

    # Total vehicles by status
    r_vehicles = await db.execute(
        select(Vehicle.status, func.count(Vehicle.id))
        .group_by(Vehicle.status)
        .order_by(Vehicle.status)
    )
    vehicle_counts = r_vehicles.all()
    total_vehicles = sum(c for _, c in vehicle_counts)

    # Orders in range
    order_stmt = select(
        func.count(CapturedOrder.id).label("total_orders"),
        func.coalesce(func.sum(CapturedOrder.amount), Decimal("0.000")).label("total_amount"),
    )
    if date_from or date_to:
        order_stmt = _apply_datetime_date_filter(order_stmt, CapturedOrder.captured_at, date_from, date_to)
    r_orders = await db.execute(order_stmt)
    order_row = r_orders.one()
    total_orders = order_row.total_orders or 0
    total_order_amount = order_row.total_amount or Decimal("0.000")

    # Attendance in range
    att_stmt = select(
        func.count(AttendanceRecord.id).label("total"),
        func.sum(
            case((AttendanceRecord.status == "present", 1), else_=0)
        ).label("present"),
        func.sum(
            case((AttendanceRecord.status == "late", 1), else_=0)
        ).label("late"),
        func.sum(
            case((AttendanceRecord.status == "absent", 1), else_=0)
        ).label("absent"),
    )
    att_stmt = _apply_date_filter(att_stmt, AttendanceRecord.date, date_from, date_to)
    r_att = await db.execute(att_stmt)
    att_row = r_att.one()
    att_total = att_row.total or 0
    att_present = att_row.present or 0
    att_late = att_row.late or 0
    att_absent = att_row.absent or 0
    att_rate = (
        f"{((att_present + att_late) / att_total * 100):.1f}%"
        if att_total > 0
        else "N/A"
    )

    # Maintenance cost in range
    maint_stmt = select(
        func.count(MaintenanceRecord.id).label("total"),
        func.coalesce(func.sum(MaintenanceRecord.cost), Decimal("0.000")).label("total_cost"),
    )
    maint_stmt = _apply_date_filter(maint_stmt, MaintenanceRecord.date, date_from, date_to)
    r_maint = await db.execute(maint_stmt)
    maint_row = r_maint.one()
    total_maint = maint_row.total or 0
    total_maint_cost = maint_row.total_cost or Decimal("0.000")

    # Cash summary in range
    cash_stmt = select(
        CashRecord.record_type,
        func.coalesce(func.sum(CashRecord.amount), Decimal("0.000")).label("total"),
    ).group_by(CashRecord.record_type)
    cash_stmt = _apply_date_filter(cash_stmt, CashRecord.date, date_from, date_to)
    r_cash = await db.execute(cash_stmt)
    cash_by_type: dict[str, Decimal] = {row[0]: row[1] for row in r_cash.all()}

    # Build summary table
    date_range_str = _date_range_label(date_from, date_to)
    headers = ["Metric", "Value"]
    rows: list[list] = [
        ["Report Period", date_range_str],
        ["", ""],
        ["--- DRIVERS ---", ""],
        ["Total Active Drivers", str(total_drivers)],
    ]
    for platform, count in platform_counts:
        rows.append([f"  Platform: {platform or 'unassigned'}", str(count)])

    rows += [
        ["", ""],
        ["--- VEHICLES ---", ""],
        ["Total Vehicles", str(total_vehicles)],
    ]
    for vstatus, count in vehicle_counts:
        rows.append([f"  Status: {vstatus}", str(count)])

    rows += [
        ["", ""],
        ["--- ORDERS ---", ""],
        ["Total Orders", str(total_orders)],
        ["Total Order Amount (KWD)", _fmt_kwd(total_order_amount)],
        ["", ""],
        ["--- ATTENDANCE ---", ""],
        ["Total Attendance Records", str(att_total)],
        ["Present", str(att_present)],
        ["Late", str(att_late)],
        ["Absent", str(att_absent)],
        ["Attendance Rate (present+late)", att_rate],
        ["", ""],
        ["--- MAINTENANCE ---", ""],
        ["Total Maintenance Records", str(total_maint)],
        ["Total Maintenance Cost (KWD)", _fmt_kwd(total_maint_cost)],
        ["", ""],
        ["--- CASH ---", ""],
    ]
    for ctype, ctotal in sorted(cash_by_type.items()):
        rows.append([f"  {ctype.capitalize()} (KWD)", _fmt_kwd(ctotal)])

    return headers, rows, "Fleet Overview Report"


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------


def _render_pdf(title: str, headers: list[str], rows: list[list]) -> bytes:
    """Render report as a PDF using reportlab.

    Uses standard Helvetica throughout (Arabic RTL rendering requires
    additional font embedding — left for a future enhancement).
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise ImportError(
            "reportlab is required for PDF generation. "
            "Install it with: pip install reportlab"
        ) from exc

    buffer = io.BytesIO()

    # Use landscape for wide reports (many columns)
    pagesize = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = styles["Title"]
    story.append(Paragraph(title, title_style))

    # Generated-at timestamp
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story.append(Paragraph(f"Generated: {generated_at}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    if not rows:
        story.append(Paragraph("No data found for the selected filters.", styles["Normal"]))
        doc.build(story)
        return buffer.getvalue()

    # Table data: header row + data rows
    table_data = [headers] + rows
    col_count = len(headers)

    # Distribute column widths evenly across the available page width
    page_width = pagesize[0] - 3 * cm  # left + right margins
    col_width = page_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(
        TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F2B46")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            # Data rows
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            # Alternating row backgrounds
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E0")),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#2563EB")),
        ])
    )

    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _render_excel(title: str, headers: list[str], rows: list[list]) -> bytes:
    """Render report as an Excel workbook using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel generation. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit is 31 chars

    # --- Style constants ---
    PRIMARY_COLOR = "0F2B46"   # dark navy (FleetPulse primary)
    ACCENT_COLOR = "2563EB"    # blue accent
    ALT_ROW_COLOR = "F0F4F8"   # light blue-grey for alternating rows

    header_fill = PatternFill("solid", fgColor=PRIMARY_COLOR)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Calibri", bold=True, color=PRIMARY_COLOR, size=13)
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    current_row = 1

    # Title row
    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=max(len(headers), 1),
    )
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align
    current_row += 1

    # Generated-at row
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=max(len(headers), 1),
    )
    gen_cell = ws.cell(row=current_row, column=1, value=f"Generated: {generated_at}")
    gen_cell.font = Font(name="Calibri", italic=True, size=9, color="6B7280")
    gen_cell.alignment = center_align
    current_row += 2  # blank row between title and header

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Data rows
    for row_idx, row_data in enumerate(rows):
        is_alt = row_idx % 2 == 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = left_align
            if is_alt:
                cell.fill = alt_fill
        current_row += 1

    # Auto-fit column widths (heuristic: max content width, capped at 50)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                cell_len = len(str(row_data[col_idx - 1]))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze panes below header row (row 4 = data start after title/generated/blank/header)
    ws.freeze_panes = ws.cell(row=current_row - len(rows), column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _render_csv(headers: list[str], rows: list[list]) -> bytes:
    """Render report as UTF-8 CSV with BOM for Excel compatibility."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows(rows)
    # Prepend UTF-8 BOM so Excel auto-detects encoding correctly
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _apply_date_filter(stmt, date_col, date_from: date | None, date_to: date | None):
    """Apply an inclusive date range filter to a SQLAlchemy statement using a Date column."""
    if date_from:
        stmt = stmt.where(date_col >= date_from)
    if date_to:
        stmt = stmt.where(date_col <= date_to)
    return stmt


def _apply_datetime_date_filter(stmt, datetime_col, date_from: date | None, date_to: date | None):
    """Apply an inclusive date range filter using a DateTime column (cast to date)."""
    if date_from:
        stmt = stmt.where(func.date(datetime_col) >= date_from)
    if date_to:
        stmt = stmt.where(func.date(datetime_col) <= date_to)
    return stmt


def _fmt_kwd(value: Decimal | float | None) -> str:
    """Format a monetary value as KWD with 3 decimal places."""
    if value is None:
        return "0.000"
    try:
        return f"{Decimal(str(value)):.{KWD_PLACES}f}"
    except Exception:
        return str(value)


def _fmt_datetime(dt: datetime | None) -> str:
    """Format a datetime to a readable UTC string."""
    if dt is None:
        return ""
    if dt.tzinfo is None:
        return dt.strftime("%Y-%m-%d %H:%M")
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _date_range_label(date_from: date | None, date_to: date | None) -> str:
    """Return a human-readable date range label."""
    if date_from and date_to:
        return f"{date_from} to {date_to}"
    if date_from:
        return f"From {date_from}"
    if date_to:
        return f"Up to {date_to}"
    return "All time"
